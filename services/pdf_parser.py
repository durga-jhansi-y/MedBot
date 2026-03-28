import pdfplumber
import anthropic
import json
import os
from dotenv import load_dotenv
import docx
import openpyxl

load_dotenv()

def extract_text(file_path: str) -> str:
    ext = os.path.splitext(file_path)[1].lower()
    
    if ext == ".txt":
        with open(file_path, "r") as f:
            return f.read()
    
    elif ext == ".pdf":
        raw_text = ""
        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    raw_text += text
        return raw_text
    
    elif ext == ".docx":
        doc = docx.Document(file_path)
        return "\n".join([para.text for para in doc.paragraphs])
    
    elif ext == ".xlsx":
        wb = openpyxl.load_workbook(file_path)
        raw_text = ""
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(values_only=True):
                row_text = " ".join([str(cell) for cell in row if cell is not None])
                if row_text.strip():
                    raw_text += row_text + "\n"
        return raw_text
    
    else:
        return ""

def parse_prescription(file_path: str) -> dict:
    # Step 1: extract text based on file type
    raw_text = extract_text(file_path)
    print("Raw text extracted:", raw_text[:200])

    if not raw_text.strip():
        return {"error": "Could not extract text from file"}

    # Step 2: send to Claude to extract structured data
    client = anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))
    response = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=1000,
        system="Extract prescription info as JSON only. Fields: patient_name, medications (name, dosage, frequency, duration), doctor_name, next_appointment. Return JSON only, no extra text.",
        messages=[{"role": "user", "content": raw_text}]
    )

    print("Claude response:", response.content[0].text)

    # Step 3: clean and parse response
    response_text = response.content[0].text.strip()
    
    if response_text.startswith("```"):
        response_text = response_text.split("```")[1]
        if response_text.startswith("json"):
            response_text = response_text[4:]
    
    return json.loads(response_text)